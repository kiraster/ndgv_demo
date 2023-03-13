# from http.client import HTTPResponse
# from urllib.request import HTTPRedirectHandler
from django.http import JsonResponse, StreamingHttpResponse
from django.shortcuts import render, redirect

# from django.http import JsonResponse
# from django.core import serializers

from device_app import models
from django.views.generic.base import View
from device_app import snmp_polling

# from ndg_app_01.netmikogetinfo import NetOps
from device_app.snmp_polling import Polling
# import ndg_app_01.snmp_polling
from multiprocessing.pool import ThreadPool

from ndgv1.settings import BASE_DIR
import os

# Create your views here.


# 首页视图
class IndexView(View):

    def get(self, request):

        # 设备概览
        device_count_switch_s = models.DeviceDetail.objects.filter(
            device_model__contains='H3C S').count()
        device_count_switch_ie = models.DeviceDetail.objects.filter(
            device_model__contains='H3C IE').count()
        device_count_router_msr = models.DeviceDetail.objects.filter(
            device_model__contains='MSR').count()
        device_count_router_wx = models.DeviceDetail.objects.filter(
            device_model__contains='WX').count()
        device_count_security_f = models.DeviceDetail.objects.filter(
            device_model__contains='H3C F').count()
        device_count_security_secpath = models.DeviceDetail.objects.filter(
            device_model__contains='SecPath').count()
        # 网络设备统计
        device_network_count = device_count_switch_s + device_count_switch_ie + device_count_router_msr + device_count_router_wx
        # 安全设备统计
        device_sec_count = device_count_security_f + device_count_security_secpath
        # device_count_server = models.DeviceDetail.objects.filter(
        #     device_model__contains='Server').count()
        # device_count_all = models.DeviceDetail.objects.all().count()
        device_count_other = models.DeviceDetail.objects.exclude(
            device_model__contains='H3C S').exclude(
                device_model__contains='IE').exclude(
                    device_model__contains='MSR').exclude(
                        device_model__contains='WX').exclude(
                            device_model__contains='H3C F').exclude(
                                device_model__contains='SecPath').exclude(
                                    device_model__contains='Server').count()

        # 设备监控概览 - 饼状图
        device_health_online = models.Device.objects.filter(
            device_state_id=1).count()
        device_health_all = models.Device.objects.all().count()
        device_health_offline = device_health_all - device_health_online

        # CPU利用率概览 - 饼状图
        device_cpu_low = models.DeviceDetail.objects.filter(
            device_cpu__lte=33).count()
        device_cpu_middle = models.DeviceDetail.objects.filter(
            device_cpu__range=[33, 66]).count()
        device_cpu_high = models.DeviceDetail.objects.filter(
            device_cpu__gte=66).count()

        # 内存利用率概览 - 饼状图
        device_mem_low = models.DeviceDetail.objects.filter(
            device_mem__lt=33).count()
        device_mem_middle = models.DeviceDetail.objects.filter(
            device_mem__range=[33, 66]).count()
        device_mem_high = models.DeviceDetail.objects.filter(
            device_mem__gt=66).count()

        # CPU TOP-10
        device_cpu_orderby = models.DeviceDetail.objects.order_by(
            '-device_cpu')[:10]
        # print(type(device_cpu_orderby))

        # Mem TOP-10
        device_mem_orderby = models.DeviceDetail.objects.order_by(
            '-device_mem')[:10]
        # print(device_mem_orderby)

        return render(request, 'home.html', locals())


# 设备页面
class DeviceView(View):

    # 返回页面
    def get(self, request):
        return render(request, 'device_list.html')


# 设备列表
class DeviceListView(View):

    # 方式1 直接传递全部数据到前端 循环输出 显示
    # def get(self, request):
    #     # 先查数据
    #     device_queryset = models.Device.objects.all()
    #     # print(device_queryset.all())
    #     return render(request, 'device_list.html', locals())

    # 方式2 ajax传递 json 数据 处理
    def get(self, request):
        # 先查数据
        device_queryset = models.Device.objects.all()
        # print(device_queryset.all())
        # 定义列表装数据，预期格式{'data':[{k1:v1, k2:v2, ...},{k1:v11, k2:v2, ...}]}
        data_list = []
        for device_obj in device_queryset:
            device_dict = {
                'id': device_obj.pk,
                'sysname': device_obj.sysname,
                'device_alias': device_obj.device_detail.device_alias,
                'mgmt_ip': device_obj.mgmt_ip,
                'device_model': device_obj.device_detail.device_model,
                'device_type': device_obj.device_detail.device_type,
                'net_area': device_obj.device_detail.net_area,
                'device_location': device_obj.device_location.device_location,
                'icmp_state': device_obj.device_state.icmp_state,
                'snmp_state': device_obj.device_state.snmp_state,
                'device_cpu': device_obj.device_detail.device_cpu,
                'device_mem': device_obj.device_detail.device_mem,
                'device_sn': device_obj.device_detail.device_sn,
                'device_uptime': device_obj.device_detail.device_uptime,
                'device_version': device_obj.device_detail.device_version,
                'device_fan_state': device_obj.device_detail.device_fan_state,
                'device_power_state':
                device_obj.device_detail.device_power_state,
                'device_temperature':
                device_obj.device_detail.device_temperature,
                'community_data': device_obj.device_detail.community_data,
                'add_time': device_obj.device_detail.add_time,
                'last_sync_time': device_obj.device_detail.last_sync_time
            }
            data_list.append(device_dict)
        # 再套一层字典
        data_dict = {}
        data_dict['data'] = data_list
        return JsonResponse(data_dict)

    # 检查团体字 input 框内容是否全由空格组成
    def check_input_comm(self, input_comm):
        if input_comm.isspace() or input_comm == '' or input_comm is None:
            return True
        else:
            return False

    def post(self, request):
        device_id = request.POST.get('device_id')
        button_val = request.POST.get('button_val')
        community_data = request.POST.get('community_data')
        device_alias = request.POST.get('device_alias')

        # 检查 请求
        if button_val == 'device_modify':
            # 查询管理IP，渲染到第一个input框内的属性值
            device_obj = models.Device.objects.get(pk=device_id)
            back_dict = {'status_code': 1111, 'res': device_obj.mgmt_ip}
            return JsonResponse(back_dict)
        if button_val == 'edit_confirm':

            try:
                # 检查 是否输入了团体字
                if self.check_input_comm(community_data):
                    back_dict = {'status_code': 3333}
                    return JsonResponse(back_dict)
                else:
                    # 构造数据写入or更新
                    detail_data = {
                        'device_alias': device_alias,
                        'community_data': community_data
                    }
                    device_obj = models.Device.objects.get(pk=device_id)
                    devicedetail_id = device_obj.device_detail.pk
                    obj, create = models.DeviceDetail.objects.update_or_create(
                        defaults=detail_data, pk=devicedetail_id)

                    if create:
                        print('创建了一行数据！，不可能')
                        print('永远不会打印这一行在控制台_ ^_^ _')
                    else:
                        # print('成功执行了写入数据库')
                        back_dict = {'status_code': 1111}
                        return JsonResponse(back_dict)
            except Exception:
                back_dict = {'status_code': 4444}
                return JsonResponse(back_dict)


# 设备详情
class DeviceDetailView(View):

    def get(self, request, device_id):
        device_queryset = models.Device.objects.filter(pk=device_id)
        return render(request, 'device_detail.html', locals())

    def post(self, request, device_id):

        # 接收post过来的数据
        mgmt_ip = request.POST.get('mgmt_ip')
        username = request.POST.get('username')
        password = request.POST.get('password')
        display_cmd = request.POST.get('display_cmd')
        # print(mgmt_ip)
        # print(display_cmd)

        # 调用netmiko_get
        from device_app.netmiko_get import NetOps
        query_obj = NetOps()
        res = query_obj.get_info(mgmt_ip, username, password, display_cmd)

        # 构造字典返回给前端ajax
        back_dict = {'status_code': 1111, 'res': res}
        return JsonResponse(back_dict)


# 设备添加
class DeviceAddView(View):

    # 导入ipaddress模块，校验传入的数据是否合规的IPv4 地址
    def verify_ipaddress(self, ip):
        import ipaddress
        try:
            ipaddress.IPv4Network(ip)
            return True
        except ValueError:
            return False

    # 检查设备IP地址 input 框是否IPv4 地址
    def check_intpu_address(self, input_ip):
        validated_res = self.verify_ipaddress(input_ip)
        if validated_res:
            return True
        else:
            return False

    # 检查团体字 input 框内容是否全由空格组成
    def check_input_comm(self, input_comm):
        if input_comm.isspace() or input_comm == '' or input_comm is None:
            return True
        else:
            return False

    def get(self, request):
        # return render(request, 'device_add.html', locals())
        return render(request, 'device_add.html')

    def post(self, request):

        # mgmt_ip = '6.6.6.6.6'
        # community_data = ''
        button_val = request.POST.get('button')
        mgmt_ip = request.POST.get('mgmt_ip')
        community_data = request.POST.get('community_data')
        # print(mgmt_ip, community_data)

        # input检查 请求
        if button_val == 'check_ip':
            if not self.check_intpu_address(mgmt_ip):
                back_dict = {'status_code': 2222}
                return JsonResponse(back_dict)
        if button_val == 'check_comm':
            if self.check_input_comm(community_data):
                back_dict = {'status_code': 3333}
                return JsonResponse(back_dict)

        # snmp测试 请求
        if button_val == 'snmp_test':

            if not self.check_intpu_address(mgmt_ip):
                back_dict = {'status_code': 2222}
                return JsonResponse(back_dict)
            if self.check_input_comm(community_data):
                back_dict = {'status_code': 3333}
                return JsonResponse(back_dict)
            else:
                sync_obj = Polling()
                res = sync_obj.do_pysnmp(mgmt_ip, community_data)
                # print(res, type(res))
                if res:
                    back_dict = {'status_code': 1111}
                    return JsonResponse(back_dict)
                else:
                    back_dict = {'status_code': 4444}
                    return JsonResponse(back_dict)

        # 继续添加 请求 和 添加并返回设备列表 请求
        if button_val == 'add_continue' or button_val == 'add_device':
            try:
                if not self.check_intpu_address(mgmt_ip):
                    back_dict = {'status_code': 2222}
                    return JsonResponse(back_dict)
                if self.check_input_comm(community_data):
                    back_dict = {'status_code': 3333}
                    return JsonResponse(back_dict)
                else:
                    device_alias = request.POST.get('device_alias')
                    # devicedetail 表写入
                    models.DeviceDetail.objects.create(
                        mgmt_ip=mgmt_ip,
                        device_alias=device_alias,
                        community_data=community_data)
                    # device 表写入，外键添加
                    devicedetail_obj = models.DeviceDetail.objects.get(
                        mgmt_ip=mgmt_ip)
                    # print(devicedetail_obj, devicedetail_obj.pk)
                    models.Device.objects.create(
                        mgmt_ip=mgmt_ip, device_detail_id=devicedetail_obj.pk)
                    # # 跳转到设备列表页面
                    # return redirect('device')
                    back_dict = {'status_code': 1111}
                    return JsonResponse(back_dict)
            except Exception:
                back_dict = {'status_code': 4444}
                return JsonResponse(back_dict)

        back_dict = {'status_code': 1111}
        return JsonResponse(back_dict)


# 设备修改
class DeviceModifyView(View):

    def get(self, request):
        pass


# 设备删除
class DeviceDeleteView(View):

    # 无ajax 直接删除
    # def get(self, request, device_id):
    #     print(device_id)
    #     del_obj = models.Device.objects.get(pk=device_id)
    #     del_id = del_obj.device_detail_id
    #     # print(del_pk)
    #     # devicedetail_id = del_obj.device_detail
    #     # print(devicedetail_id)
    #     del_obj = models.DeviceDetail.objects.filter(pk=del_id)
    #     if del_obj.exists():
    #         # do_deltele = del_obj.last()
    #         del_obj.delete()
    # # 获取对象
    # del_obj = models.Device.objects.filter(pk=device_id)
    # devicedetail_id = del_obj.device_detail_id
    # print(devicedetail_id)
    # models.DeviceDetail.objects.filter(pk=del_obj.device).delete()
    # del_obj.delete()
    # 跳转设备列表页面
    # return redirect('device_list')
    # def get(self, request, device_id):
    #     # back_dict = {'code': 1000, 'msg': ''}
    #     # device_id = request.POST.get('device_id')
    #     # models.Device.objects.filter(pk=device_id).delete()
    #     # back_dict['msg'] = '删除已完成！'
    #     # return JsonResponse(back_dict)
    #     # device_id = request.POST.get('device_id')
    #     query_res = models.Device.objects.filter(pk=device_id)
    #     if not query_res:
    #         return HTTPResponse(json.dumps({'state': False, 'erro': 'Date Exists.'}))
    #     models.Device.objects.filter(pk=device_id).delete()
    #     models.DeviceDetail.objects.filter(pk=device_id).delete()
    #     return HTTPResponse(json.dumps({'state': True}))

    # 无ajax 直接删除
    # def get(self, request, device_id):
    #     # print(device_id)
    #     del_obj = models.Device.objects.get(pk=device_id)
    #     del_id = del_obj.device_detail_id
    #     del_obj = models.DeviceDetail.objects.filter(pk=del_id)
    #     if del_obj.exists():
    #         del_obj.delete()
    #     return redirect('device')

    def post(self, request):
        # if request.is_ajax():
        if request.method == 'POST':
            try:
                back_dict = {'status_code': 1111}
                device_id = request.POST.get('device_id')
                del_obj = models.Device.objects.get(pk=device_id)
                del_id = del_obj.device_detail_id
                del_obj = models.DeviceDetail.objects.filter(pk=del_id)
                if not del_obj:
                    back_dict['status_code'] = 2222
                    return JsonResponse(back_dict)
                del_obj.delete()
                # back_dict['msg'] = '数据已删除'
                return JsonResponse(back_dict)
            except Exception:
                back_dict['status_code'] = 2222
                return JsonResponse(back_dict)


# 设备同步
class DeviceSyncView(View):

    def post(self, request):

        # 获取post 过来的device_id
        device_id = request.POST.get('device_id')
        # print(device_id)

        # 根据device_id 查找 管理IP，snmp团体字
        device_obj = models.Device.objects.get(pk=device_id)
        mgmt_ip = device_obj.device_detail.mgmt_ip
        community_data = device_obj.device_detail.community_data
        # print(mgmt_ip, community_data)

        # 调用ops脚本 进行 icmp 检查， 传递给写入数据库函数， 对应的icmp状态
        # 调用轮询脚本，接收返回的字典结果
        # 调用snmp_polling ，实例化后传入参数，接收返回结果
        sync_obj = Polling()
        device_data, detail_data, devicelocation_data = sync_obj.run(
            mgmt_ip, community_data)

        # 调用函数create_date 写入数据库
        flag = self.create_date(device_id, device_obj, device_data,
                                detail_data, devicelocation_data)
        back_dict = {'status_code': 1111}
        if not flag:
            back_dict['status_code'] = 2222
            return JsonResponse(back_dict)
        return JsonResponse(back_dict)

    def create_date(self, device_id, device_obj, device_data, detail_data,
                    devicelocation_data):

        try:
            # 方式1，传统的 update，由于导入前已有数据，不存在找不到主键pk，和没有创建的情况
            # devicedetail 表写入
            # models.DeviceDetail.objects.filter(pk=device_id).update(
            #     device_uptime=data['sysuptime'],
            #     device_version=data['sysversion'],
            #     device_sn=data['sn'],
            #     device_model=data['model'],
            #     device_cpu=data['cpu_utilization'],
            #     device_mem=data['mem_utilization'],
            #     device_power_state=data['power_state'],
            #     device_fan_state=data['fan_state'],
            #     device_temperature=data['sys_temperature'],
            # )

            # # device 表写入
            # models.Device.objects.filter(pk=device_id).update(sysname=data['sysname'])

            # 方式２，　update_or_create
            # 先查先写devicelocation 表，
            location_data = devicelocation_data['device_location']
            # print(location_data)
            location_obj = models.DeviceLocation.objects.filter(
                device_location=location_data).first()
            if not location_obj:
                devicelocation_obj = models.DeviceLocation.objects.create(
                    device_location=location_data)
                device_location_id = devicelocation_obj
                # print(device_location_id)
            else:
                device_location_id = location_obj.pk
                # print(device_location_id)

            device_data['device_location_id'] = device_location_id
            devicedetail_id = device_obj.device_detail.pk
            # device_data写入device表, detail_data写入devicedetail表

            obj1, create1 = models.Device.objects.update_or_create(
                defaults=device_data, pk=device_id)
            obj2, create2 = models.DeviceDetail.objects.update_or_create(
                defaults=detail_data, pk=devicedetail_id)

            if create1 and create2:
                print('创建了一行数据！，不可能')
                print('永远不会打印这一行在控制台_ ^_^ _')
            else:
                # print('成功执行了写入数据库')
                pass
            return True
        except Exception as e:
            print(str(e))
            return False


# 设备数据上传
class DeviceUploadView(View):

    def post(self, request):
        file_obj = request.FILES.get('file')
        file_name = file_obj.name
        # print(file_name)
        # 文件路径

        file_path = os.path.join(BASE_DIR, 'upload_files', file_name)
        # print(file_path)
        try:
            try:
                info_dict = {}
                with open(file_path, 'wb+') as f:
                    for chunk in file_obj.chunks():
                        f.write(chunk)
            except Exception as e:
                info_dict['res1'] = '上传错误1' + str(e)
                return JsonResponse(info_dict)
            info_dict['res1'] = '上传成功'
            res = self.read_date(file_path)
            # print(type(res))
            if isinstance(res, str):
                info_dict['res2'] = res
                return JsonResponse(info_dict)
            flag, res = self.create_date(res)
            if not flag:
                info_dict['res2'] = res
                return JsonResponse(info_dict)

            info_dict['res2'] = '导入成功'
            return JsonResponse(info_dict)
        except Exception as e:
            print(str(e))
            info_dict['res1'] = '上传错误2 ' + str(e)
            return JsonResponse(info_dict)

    def read_date(self, date_source):

        try:
            # 载入数据表格
            from openpyxl import load_workbook
            wb = load_workbook(date_source, read_only=True)

            # 读取上传的数据表格
            ws1 = wb[wb.sheetnames[0]]
            # 遍历数据行， 指定范围
            for row in ws1.iter_rows(min_row=2, max_col=19):
                data_dict = {
                    'sysname': row[1].value,
                    'device_alias': row[2].value,
                    'mgmt_ip': row[3].value,
                    'net_area': row[4].value,
                    'device_location': row[5].value,
                    'device_type': row[6].value,
                    'device_model': row[7].value,
                    'device_version': row[8].value,
                    'device_sn': row[9].value,
                    'community_data': row[10].value,
                    'snmp_state': row[11].value,
                    'icmp_state': row[12].value,
                    'device_uptime': row[13].value,
                    'device_cpu': row[14].value,
                    'device_mem': row[15].value,
                    'device_power_state': row[16].value,
                    'device_fan_state': row[17].value,
                    'device_temperature': row[18].value,
                }
                yield data_dict
            wb.close()
        except Exception as e:
            return str(e)
        finally:
            wb.close()

    def create_date(self, data):
        try:
            # 写入数据库
            for data_dict in data:

                # devicedetail 表写入
                models.DeviceDetail.objects.create(
                    mgmt_ip=data_dict['mgmt_ip'],
                    device_alias=data_dict['device_alias'],
                    net_area=data_dict['net_area'],
                    device_type=data_dict['device_type'],
                    device_model=data_dict['device_model'],
                    device_version=data_dict['device_version'],
                    device_sn=data_dict['device_sn'],
                    community_data=data_dict['community_data'],
                )
                # devicelocation 表写入
                # print(data_dict, type(data_dict))
                # print(data_dict['device_location'])
                if data_dict['device_location'] is not None:
                    models.DeviceLocation.objects.create(
                        device_location=data_dict['device_location'])
                    devicedetail_obj = models.DeviceDetail.objects.filter(
                        mgmt_ip=data_dict['mgmt_ip']).first()
                    devicelocation_obj = models.DeviceLocation.objects.filter(
                        device_location=data_dict['device_location']).first()
                    models.Device.objects.create(
                        mgmt_ip=data_dict['mgmt_ip'],
                        device_detail_id=devicedetail_obj.pk,
                        device_location_id=devicelocation_obj.pk)

                else:

                    # device 表写入
                    devicedetail_obj = models.DeviceDetail.objects.filter(
                        mgmt_ip=data_dict['mgmt_ip']).first()
                    models.Device.objects.create(
                        mgmt_ip=data_dict['mgmt_ip'],
                        device_detail_id=devicedetail_obj.pk)

                # # 查 devicelocation表，判断，写入device表 外键
                # devicelocation_obj = models.DeviceLocation.objects.filter(
                #     device_location=data_dict['device_location']).first()

                # if not devicelocation_obj:
                #     models.Device.objects.create(
                #         device_location_id=devicelocation_obj.pk, )

                # # 跳转到设备列表页面
                # return redirect('device')
            return True, 'Success, Compelete.'
        except Exception as e:
            return False, str(e) + '-导入的数据与数据库内数据冲突'


# 上传模板下载
class DownloadView(View):

    def get(self, request):

        file_path = os.path.join(BASE_DIR, 'static', 'upload_template',
                                 'IMPORT_TEMPLATE.xlsx')

        # print(file_path)

        # file_path = os.path.join(BASE_DIR, 'static', 'upload_template',
        #                          '设备导入模版.xlsx')

        # do something...
        def down_chunk_file_manager(file_path, chuck_size=1024):
            with open(file_path, "rb") as file:
                while True:
                    chuck_stream = file.read(chuck_size)
                    if chuck_stream:
                        yield chuck_stream
                    else:
                        break

        response = StreamingHttpResponse(down_chunk_file_manager(file_path))
        response['Content-Type'] = 'application/octet-stream'
        # 纯英文文件名
        response['Content-Disposition'] = 'attachment;filename="{0}"'.format(
            'IMPORT_TEMPLATE.xlsx')
        # 带中文文件名
        # from django.utils.encoding import escape_uri_path
        # response[
        #     "Content-Disposition"] = "attachment; filename*=UTF-8''{}".format(
        #         escape_uri_path('设备导入模版.xlsx'))

        return response


# 数据库表数据全部删除
class DeleteAnythingView(View):

    def get(self, request):
        import time

        devicedetail_obj = models.DeviceDetail.objects.all()
        devicedetail_obj.delete()
        time.sleep(10)
        devicelocation_obj = models.DeviceLocation.objects.all()
        devicelocation_obj.delete()
        time.sleep(5)
        device_obj = models.Device.objects.all()
        device_obj.delete()
        time.sleep(5)
        return redirect('device')

    def post(self, request):
        # if request.is_ajax():
        # if request.method == 'POST':
        try:
            back_dict = {'status_code': 1111}

            # 只直接使用原生 sql 语句 更快
            from django.db import connection
            # sqlite 数据库
            sqlite_db = connection.cursor()
            # MySQL 数据库
            # mysql_db = connection.cursor()
            # 执行 命令 重置 自增ID
            sqlite_db.execute('DELETE FROM sqlite_sequence')
            sqlite_db.execute('DELETE FROM device_app_device')
            sqlite_db.execute('DELETE FROM device_app_devicedetail')
            sqlite_db.execute('DELETE FROM device_app_devicestate')
            sqlite_db.execute('DELETE FROM device_app_devicelocation')
            sqlite_db.execute('DELETE FROM scheduler_app_schedulerdetail')
            sqlite_db.execute('DELETE FROM django_apscheduler_djangojob')
            sqlite_db.execute('DELETE FROM django_apscheduler_djangojobexecution')
            sqlite_db.execute('DELETE FROM sqlite_sequence')
            # 取消外键约束
            # mysql_db.execute('SET FOREIGN_KEY_CHECKS=0')
            # mysql_db.execute('truncate table device_app_device')
            # mysql_db.execute('truncate table device_app_devicedetail')
            # mysql_db.execute('truncate table device_app_devicestate')
            # mysql_db.execute('truncate table device_app_devicelocation')
            # mysql_db.execute('truncate table scheduler_app_schedulerdetail')
            # mysql_db.execute('truncate table django_apscheduler_djangojob')
            # mysql_db.execute('truncate table django_apscheduler_djangojobexecution')
            # 设置外键约束
            # mysql_db.execute('SET FOREIGN_KEY_CHECKS=1')
            # 初始化一些数据
            # 设备状态表初始化
            data_list = [models.DeviceState(snmp_state=True, icmp_state=True), models.DeviceState(snmp_state=False, icmp_state=False), models.DeviceState(snmp_state=True, icmp_state=False), models.DeviceState(snmp_state=False, icmp_state=True)]
            models.DeviceState.objects.bulk_create(data_list)
            # 设备位置表初始化
            models.DeviceLocation.objects.create(device_location='信息机房')

            return JsonResponse(back_dict)
        except Exception as e:
            back_dict['status_code'] = 2222
            print(e)
            print('清空出错2222')
            return JsonResponse(back_dict)


# 同步所有-->>> 已转至 计划任务界面， 创建任务后 ， 后台轮询snmp 和 写入
class SyncAnythingView(View):

    def scq(self):
        # queryset对象，管理IP，snmp团体字
        devicedetail_obj = models.DeviceDetail.objects.all()
        for row_obj in devicedetail_obj:
            polling_dict = {
                'mgmt_ip': row_obj.mgmt_ip,
                'community_data': row_obj.community_data,
            }
            yield polling_dict

    def get(self, request):

        # 获取生成器
        polling_work = self.scq()
        # print(type(polling_work))
        # 异步执行获取写入的数据
        # ThreadPool 设定异步进程数为 x
        pool = ThreadPool(512)
        # sync_obj = Polling()
        for row_dict in polling_work:
            # device_data, detail_data = sync_obj.run(mgmt_ip, community_data)
            res = pool.apply_async(snmp_polling.run, args=(row_dict, ))
            # print(res.get())
            # 调用函数create_date 写入数据库
            self.create_date(row_dict['mgmt_ip'],
                             res.get()[0],
                             res.get()[1],
                             res.get()[2])

        pool.close()
        pool.join()

        back_dict = {'status_code': 1111}
        return JsonResponse(back_dict)

    def create_date(self, mgmt_ip, device_data, detail_data,
                    devicelocation_data):
        try:
            # 方式２，　update_or_create

            # 先查先写devicelocation 表，
            location_data = devicelocation_data['device_location']
            location_obj = models.DeviceLocation.objects.filter(
                device_location=location_data).first()
            if not location_obj:
                devicelocation_obj = models.DeviceLocation.objects.create(
                    device_location=location_data)
                device_location_id = devicelocation_obj
            else:
                device_location_id = location_obj.pk

            device_data['device_location_id'] = device_location_id

            # device_data写入device表, detail_data写入devicedetail表
            obj1, create1 = models.Device.objects.update_or_create(
                defaults=device_data, mgmt_ip=mgmt_ip)
            obj2, create2 = models.DeviceDetail.objects.update_or_create(
                defaults=detail_data, mgmt_ip=mgmt_ip)

            if create1 and create2:
                print('创建了一行数据！，不可能')
                print('永远不会打印这一行在控制台_ ^_^ _')
            else:
                # print('成功执行了写入数据库')
                pass
            return True
        except Exception:
            return False
